import os
import shutil
from datetime import datetime
from typing import Any

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from constants.script_consts import TEMPLATE_PATH, get_next_iteration_dir


class ExcelWriter:
    def __init__(self, template_path: str = TEMPLATE_PATH) -> None:
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template Excel introuvable : {template_path}")

        self.output_dir = get_next_iteration_dir()
        os.makedirs(self.output_dir, exist_ok=True)

        timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%Mmin%Ss")
        filename = f"resultats_{timestamp}.xlsx"
        self.output_path = os.path.join(self.output_dir, filename)

        shutil.copy(template_path, self.output_path)

        self.workbook: Workbook = load_workbook(self.output_path)

        sheet = self.workbook.active

        if sheet is None:
            raise ValueError("Aucune feuille active trouvée dans le fichier Excel.")

        self.sheet: Worksheet = sheet

        self.header_map: dict[str, int] = {
            str(cell.value): idx + 1
            for idx, cell in enumerate(self.sheet[1])
            if cell.value
        }

        self.next_row = self._first_available_row()

    def _row_has_content(self, row: int) -> bool:
        return any(cell.value not in (None, "") for cell in self.sheet[row])

    def _first_available_row(self) -> int:
        row = 2

        while self._row_has_content(row):
            row += 1

        return row

    def _format_answer(self, answer: Any) -> str:
        if isinstance(answer, list):
            return "; ".join(str(item) for item in answer)

        if answer is None:
            return ""

        return str(answer)

    def insert_row(self, pdf_name: str, responses: dict[str, Any]) -> None:
        row = self.next_row

        self.sheet.cell(row=row, column=1, value=pdf_name)
        self.sheet.cell(
            row=row,
            column=2,
            value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

        for qid, answer in responses.items():
            col_idx = self.header_map.get(qid)

            if not col_idx:
                continue

            self.sheet.cell(
                row=row,
                column=col_idx,
                value=self._format_answer(answer),
            )

        self.next_row += 1

    def insert_blank_row(self) -> None:
        self.next_row += 1

    def insert_article_results(
        self,
        pdf_name: str,
        responses_en: dict[str, Any],
        responses_fr: dict[str, Any],
        partial: bool = False,
    ) -> None:
        suffix = " [PARTIAL]" if partial else ""

        self.insert_row(f"{pdf_name}{suffix} (EN)", responses_en)
        self.insert_row(f"{pdf_name}{suffix} (FR)", responses_fr)
        self.insert_blank_row()

        self.workbook.save(self.output_path)