import os
import shutil
from datetime import datetime
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from constants.script_consts import TEMPLATE_PATH, get_next_iteration_dir
from retrieval.consts import RECOVERY_OUTPUT_PREFIX


class RecoveredExcelWriter:
    def __init__(
        self,
        template_path: str = TEMPLATE_PATH,
        output_dir: str | None = None,
        output_path: str | None = None,
    ) -> None:
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template Excel introuvable : {template_path}")

        if output_path is not None:
            self.output_path = output_path
            self.output_dir = os.path.dirname(output_path) or "."
        else:
            self.output_dir = output_dir or get_next_iteration_dir()
            timestamp = datetime.now().strftime("%Y-%m-%d_%Hh%Mmin%Ss")
            filename = f"{RECOVERY_OUTPUT_PREFIX}_{timestamp}.xlsx"
            self.output_path = os.path.join(self.output_dir, filename)

        os.makedirs(self.output_dir, exist_ok=True)
        shutil.copy(template_path, self.output_path)

        self.workbook: Workbook = load_workbook(self.output_path)

        sheet = self.workbook.active
        if sheet is None:
            raise ValueError("Aucune feuille active trouvée dans le fichier Excel.")

        self.sheet: Worksheet = sheet

        self.header_map: dict[str, int] = {
            str(cell.value): idx + 1
            for idx, cell in enumerate(self.sheet[1])
            if cell.value
        }

        self.next_row = self._first_available_row()

    def _row_has_content(self, row: int) -> bool:
        return any(cell.value not in (None, "") for cell in self.sheet[row])

    def _first_available_row(self) -> int:
        row = 2

        while self._row_has_content(row):
            row += 1

        return row

    def _format_answer(self, answer: Any) -> str:
        if answer is None:
            return ""

        if isinstance(answer, list):
            return "; ".join(str(item) for item in answer)

        return str(answer)

    def _lock_file_path(self) -> str:
        dirname = os.path.dirname(self.output_path)
        filename = os.path.basename(self.output_path)

        return os.path.join(dirname, f".~lock.{filename}#")

    def _assert_not_open_in_libreoffice(self) -> None:
        lock_path = self._lock_file_path()

        if os.path.exists(lock_path):
            raise RuntimeError(
                f"Excel file appears to be open in LibreOffice: {self.output_path}. "
                "Close it before running recovery."
            )

    def insert_row(self, row_name: str, responses: dict[str, Any]) -> None:
        row = self.next_row

        self.sheet.cell(row=row, column=1, value=row_name)
        self.sheet.cell(
            row=row,
            column=2,
            value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

        for qid, answer in responses.items():
            col_idx = self.header_map.get(qid)

            if not col_idx:
                continue

            self.sheet.cell(
                row=row,
                column=col_idx,
                value=self._format_answer(answer),
            )

        self.next_row += 1

    def insert_blank_row(self) -> None:
        self.next_row += 1

    def insert_article_results(
        self,
        pdf_name: str,
        responses_en: dict[str, Any],
        responses_fr: dict[str, Any],
        partial: bool = False,
    ) -> None:
        suffix = " [PARTIAL]" if partial else ""

        self.insert_row(f"{pdf_name}{suffix} (EN)", responses_en)
        self.insert_row(f"{pdf_name}{suffix} (FR)", responses_fr)
        self.insert_blank_row()
        self.save()

        if not self.has_persisted_article(pdf_name):
            raise RuntimeError(
                "Excel save verification failed. "
                f"Article was written in memory but was not found on disk: {pdf_name}"
            )

    def save(self) -> None:
        self._assert_not_open_in_libreoffice()
        self.workbook.save(self.output_path)

    @staticmethod
    def _normalize_article_name(row_name: str | None) -> str | None:
        if not isinstance(row_name, str):
            return None

        if row_name.endswith(" (EN)") or row_name.endswith(" (FR)"):
            row_name = row_name[:-5]
        else:
            return None

        if row_name.endswith(" [PARTIAL]"):
            row_name = row_name[:-10]

        return row_name

    @classmethod
    def unique_article_names_from_path(cls, workbook_path: str) -> list[str]:
        workbook = load_workbook(workbook_path, read_only=True)
        sheet = workbook.active
        article_names: set[str] = set()

        try:
            if sheet is None:
                return []

            for row_name, *_ in sheet.iter_rows(
                min_row=2,
                max_col=1,
                values_only=True,
            ):
                normalized = cls._normalize_article_name(row_name)

                if normalized:
                    article_names.add(normalized)

            return sorted(article_names)

        finally:
            workbook.close()

    def has_persisted_article(self, article_name: str) -> bool:
        return article_name in set(
            self.unique_article_names_from_path(self.output_path)
        )

    def verify_persisted_articles(self, expected_article_names: Iterable[str]) -> None:
        expected = set(expected_article_names)
        actual = set(self.unique_article_names_from_path(self.output_path))

        if actual != expected:
            missing = sorted(expected - actual)
            extra = sorted(actual - expected)

            raise ValueError(
                "Recovered workbook integrity check failed. "
                f"expected={len(expected)}, actual={len(actual)}, "
                f"missing={missing}, extra={extra}"
            )